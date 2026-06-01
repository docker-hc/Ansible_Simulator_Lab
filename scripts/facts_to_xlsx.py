#!/usr/bin/env python3
"""facts_to_xlsx.py — workbook with diff highlighting, HBA-WWN tab, empty-HG flag."""

import argparse
import datetime as dt
import glob
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FACT_TYPES = ["system", "parity_groups", "ports", "pools", "ldevs", "host_groups", "host_mode_options", "journals", "licenses"]
FACT_LABELS = {
    "system": "Storage System", "parity_groups": "Parity Groups",
    "ports": "Storage Ports", "pools": "Storage Pools",
    "ldevs": "LDEVs", "host_groups": "Host Groups",
    "host_mode_options": "Host Mode Options",
    "journals": "Journals",
    "licenses": "Licenses",
}
KEY_FIELDS = {
    "system": ["serial_number"], "parity_groups": ["parity_group_id"],
    "ports": ["port_id"], "pools": ["pool_id", "id", "pool_name"],
    "ldevs": ["ldev_id", "id"],
    "host_groups": [("port_id", "host_group_number"), ("port_id", "host_group_name")],
    "host_mode_options": ["host_mode_option_id"],
    "journals": ["journalId"],
    "licenses": ["PRO_ID"],
}
REG_WWN_FIELDS = ["wwns", "host_wwns", "hba_wwns", "wwn_list"]

NAVY = "1F3A5F"; STEEL = "2E5E8C"; LIGHT = "E8F0F7"; ZEBRA = "F4F8FB"
WHITE = "FFFFFF"; RULE = "B8CDE0"
ADD_FILL = "C6EFCE"; ADD_TEXT = "006100"
CHG_FILL = "FFEB9C"; CHG_TEXT = "9C6500"
DEL_FILL = "FFC7CE"; DEL_TEXT = "9C0006"
EMPTY_FILL = "FFF2CC"; EMPTY_TEXT = "7F6000"

FONT = "Arial"
H_FONT = Font(name=FONT, bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name=FONT, bold=True, color=WHITE, size=14)
SECTION_FONT = Font(name=FONT, bold=True, color=WHITE, size=11)
CELL_FONT = Font(name=FONT, size=10)
LABEL_FONT = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color=RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def load_raw(raw_dir):
    arrays = {}
    for ft in FACT_TYPES:
        path = os.path.join(raw_dir, f"{ft}.json")
        if not os.path.exists(path):
            continue
        with open(path) as fh:
            for item in json.load(fh):
                serial = str(item.get("item", {}).get("serial") or item.get("serial", "unknown"))
                payload = extract_payload(item.get("ansible_facts", {}))
                if ft == "host_mode_options" and isinstance(payload, dict):
                    payload = payload.get("host_mode_options", [])
                if ft == "journals":
                    payload = payload or []
                if payload is None:
                    payload = []
                arrays.setdefault(serial, {})[ft] = payload
    hba_path = os.path.join(raw_dir, "hba_wwn.json")
    if os.path.exists(hba_path):
        with open(hba_path) as fh:
            for item in json.load(fh):
                serial = str(item.get("item", {}).get("serial") or item.get("serial", "unknown"))
                recs = item.get("ansible_facts", {}).get("hba_wwns", [])
                arrays.setdefault(serial, {}).setdefault("_hba", []).extend(recs)
    return arrays


def extract_payload(ansible_facts):
    data = {k: v for k, v in ansible_facts.items() if k != "user_consent_required"}
    if not data:
        return {}
    if len(data) == 1:
        return next(iter(data.values()))
    for v in data.values():
        if isinstance(v, list):
            return v
    return data


def as_list(payload):
    if isinstance(payload, dict):
        return [payload]
    return payload if isinstance(payload, list) else []


def row_key(fact_type, row):
    for cand in KEY_FIELDS.get(fact_type, []):
        if isinstance(cand, tuple):
            if all(c in row for c in cand):
                return "|".join(str(row.get(c, "")) for c in cand)
        elif cand in row:
            return str(row.get(cand, ""))
    return None


def flat(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, separators=(",", ":"))
    return "" if value is None else value


def extract_registered_wwns(hg_row):
    out = []
    port = hg_row.get("port_id", "")
    gid = hg_row.get("host_group_number", "")
    gname = hg_row.get("host_group_name", "")
    for fld in REG_WWN_FIELDS:
        val = hg_row.get(fld)
        if isinstance(val, list):
            for w in val:
                if isinstance(w, dict):
                    wwn = w.get("wwn") or w.get("host_wwn") or w.get("id") or ""
                    nick = w.get("nickname") or w.get("name") or w.get("wwn_nickname") or ""
                else:
                    wwn, nick = str(w), ""
                if wwn:
                    out.append({"port_id": port, "host_group_number": gid,
                                "host_group_name": gname, "hba_wwn": wwn,
                                "nickname": nick, "source": "registered"})
    return out


def all_hba_records(serial, a):
    recs = list(a.get("_hba", []))
    for hg in as_list(a.get("host_groups", [])):
        if isinstance(hg, dict):
            recs.extend(extract_registered_wwns(hg))
    return recs


def annotate_empty_hgs(serial, a):
    recs = all_hba_records(serial, a)
    seen = {}
    for r in recs:
        k = (str(r.get("port_id", "")), str(r.get("host_group_number", "")))
        seen.setdefault(k, set()).add(str(r.get("hba_wwn", "")))
    counts = {k: len(v) for k, v in seen.items()}
    for hg in as_list(a.get("host_groups", [])):
        if not isinstance(hg, dict):
            continue
        k = (str(hg.get("port_id", "")), str(hg.get("host_group_number", "")))
        n = counts.get(k, 0)
        hg["hba_count"] = n
        hg["empty"] = "Yes" if n == 0 else "No"


def style_header(ws, row, ncols, fill=STEEL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, max_w=70):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            widths[col] = min(max(widths.get(col, 10), len(str(cell.value)) + 2), max_w)
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def num(payload):
    if isinstance(payload, list):
        return len(payload)
    return 1 if isinstance(payload, dict) and payload else 0


def build_summary(wb, arrays):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:M1")
    t = ws["A1"]; t.value = "Hitachi VSP One B26 — Lab Inventory"
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = LEFT; ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:M2")
    ws["A2"].value = "Generated " + dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="5A7894")
    headers = ["Serial", "Model", "Microcode", "Address", "Total Capacity",
               "Free Capacity", "Efficiency", "Parity Grps", "Ports", "Pools",
               "LDEVs", "Host Grps", "HBA WWNs"]
    hrow = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(headers))
    ws.freeze_panes = f"A{hrow + 1}"
    r = hrow + 1
    for serial in sorted(arrays):
        a = arrays[serial]
        sysd = a.get("system", {}) if isinstance(a.get("system"), dict) else {}
        eff = sysd.get("total_efficiency", {})
        eff_ratio = eff.get("total_ratio", "") if isinstance(eff, dict) else ""
        values = [serial, sysd.get("model", ""), sysd.get("microcode_version", ""),
                  sysd.get("controller_address", "") or a.get("_addr", ""),
                  sysd.get("total_capacity", ""), sysd.get("free_capacity", ""),
                  (eff_ratio + ":1") if eff_ratio else "",
                  num(a.get("parity_groups")), num(a.get("ports")), num(a.get("pools")),
                  num(a.get("ldevs")), num(a.get("host_groups")),
                  len(all_hba_records(serial, a))]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = CELL_FONT; cell.border = BORDER
            cell.alignment = CENTER if i >= 8 else LEFT
            if (r - hrow) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
    for col in range(8, 14):
        L = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({L}{hrow+1}:{L}{r-1})")
        cell.font = LABEL_FONT; cell.alignment = CENTER; cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=LIGHT)
    autosize(ws)


def diff_arrays(cur, prev):
    changes = {}; rows = []
    if not prev:
        return changes, rows
    for serial in sorted(set(cur) | set(prev)):
        for ft in FACT_TYPES:
            c_rows = {row_key(ft, r): r for r in as_list(cur.get(serial, {}).get(ft, [])) if isinstance(r, dict)}
            p_rows = {row_key(ft, r): r for r in as_list(prev.get(serial, {}).get(ft, [])) if isinstance(r, dict)}
            c_rows.pop(None, None); p_rows.pop(None, None)
            for k in c_rows.keys() | p_rows.keys():
                c, p = c_rows.get(k), p_rows.get(k)
                if c and not p:
                    changes[(serial, ft, k)] = {"status": "added", "fields": {}}
                    rows.append([serial, FACT_LABELS[ft], k, "ADDED", ""])
                elif p and not c:
                    changes[(serial, ft, k)] = {"status": "removed", "fields": {}}
                    rows.append([serial, FACT_LABELS[ft], k, "REMOVED", ""])
                else:
                    diffs = {f: (flat(p.get(f, "")), flat(c.get(f, "")))
                             for f in set(c) | set(p)
                             if flat(c.get(f, "")) != flat(p.get(f, ""))}
                    if diffs:
                        changes[(serial, ft, k)] = {"status": "changed", "fields": diffs}
                        det = "; ".join(f"{f}: {o} -> {n}" for f, (o, n) in sorted(diffs.items()))
                        rows.append([serial, FACT_LABELS[ft], k, "CHANGED", det])
    return changes, rows


def build_changes_tab(wb, rows, label):
    ws = wb.create_sheet(title="Changes")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    t = ws["A1"]; t.value = "Changes vs previous run"
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor=NAVY); t.alignment = LEFT
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:E2"); ws["A2"].value = label
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="5A7894")
    hdr = ["Serial", "Fact Type", "Object Key", "Status", "Detail (old -> new)"]
    hrow = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(hdr)); ws.freeze_panes = f"A{hrow+1}"
    if not rows:
        ws.cell(row=hrow + 1, column=1, value="No changes since previous run.").font = \
            Font(name=FONT, italic=True, size=10, color="5A7894")
        autosize(ws); return
    fills = {"ADDED": (ADD_FILL, ADD_TEXT), "REMOVED": (DEL_FILL, DEL_TEXT),
             "CHANGED": (CHG_FILL, CHG_TEXT)}
    order = {"ADDED": 0, "CHANGED": 1, "REMOVED": 2}
    r = hrow + 1
    for row in sorted(rows, key=lambda x: (x[0], order.get(x[3], 9), x[1], x[2])):
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = LEFT
        fill, text = fills.get(row[3], (ZEBRA, "000000"))
        sc = ws.cell(row=r, column=4)
        sc.fill = PatternFill("solid", fgColor=fill)
        sc.font = Font(name=FONT, size=10, bold=True, color=text)
        r += 1
    autosize(ws)


def write_section(ws, start_row, title, payload):
    rows = as_list(payload)
    headers = []
    for row in rows:
        if isinstance(row, dict):
            for k in row:
                if k not in headers:
                    headers.append(k)
    span = max(len(headers), 1)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = SECTION_FONT; cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = LEFT
    ws.row_dimensions[start_row].height = 20
    if not headers:
        ws.cell(row=start_row + 1, column=1, value="(no data)").font = \
            Font(name=FONT, italic=True, size=9, color="5A7894")
        return start_row + 3
    hrow = start_row + 1
    for i, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(headers))
    empty_col = headers.index("empty") + 1 if "empty" in headers else None
    r = hrow + 1
    for row in rows:
        if not isinstance(row, dict):
            continue
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=i, value=flat(row.get(h, "")))
            cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = LEFT
            if (r - hrow) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        if empty_col and str(row.get("empty")) == "Yes":
            cell = ws.cell(row=r, column=empty_col)
            cell.fill = PatternFill("solid", fgColor=EMPTY_FILL)
            cell.font = Font(name=FONT, size=10, bold=True, color=EMPTY_TEXT)
        r += 1
    return r + 2


def build_detail(wb, serial, a):
    ws = wb.create_sheet(title=serial[:31])
    ws.sheet_view.showGridLines = False
    sysd = a.get("system", {}) if isinstance(a.get("system"), dict) else {}
    ws.merge_cells("A1:H1")
    t = ws["A1"]; t.value = f"Array {serial} — {sysd.get('model', 'VSP One B26')}"
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor=NAVY); t.alignment = LEFT
    ws.row_dimensions[1].height = 24
    row = 3
    for ft in FACT_TYPES:
        row = write_section(ws, row, FACT_LABELS[ft], a.get(ft, []))
    row = write_section(ws, row, "HBA WWNs (logins + registered)", all_hba_records(serial, a))
    autosize(ws)


def build_facttype_tab(wb, fact_type, arrays, changes):
    title = ("All-" + FACT_LABELS[fact_type])[:31]
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    headers, stacked = [], []
    for serial in sorted(arrays):
        for row in as_list(arrays[serial].get(fact_type, [])):
            if not isinstance(row, dict):
                continue
            for k in row:
                if k not in headers:
                    headers.append(k)
            stacked.append((serial, row))
    cols = ["serial"] + headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(cols), 1))
    t = ws.cell(row=1, column=1, value=FACT_LABELS[fact_type] + " — all arrays")
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor=NAVY); t.alignment = LEFT
    ws.row_dimensions[1].height = 24
    hrow = 3
    for i, h in enumerate(cols, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(cols)); ws.freeze_panes = f"A{hrow+1}"
    if not stacked:
        ws.cell(row=hrow + 1, column=1, value="(no data)").font = \
            Font(name=FONT, italic=True, size=9, color="5A7894")
        autosize(ws); return
    empty_col = (headers.index("empty") + 2) if "empty" in headers else None
    r = hrow + 1
    for serial, row in stacked:
        k = row_key(fact_type, row)
        ch = changes.get((serial, fact_type, k))
        ws.cell(row=r, column=1, value=serial)
        for i, h in enumerate(headers, 2):
            ws.cell(row=r, column=i, value=flat(row.get(h, "")))
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = LABEL_FONT if c == 1 else CELL_FONT
            cell.border = BORDER; cell.alignment = LEFT
            if (r - hrow) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        if ch and ch["status"] == "added":
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ADD_FILL)
                ws.cell(row=r, column=c).font = Font(name=FONT, size=10, color=ADD_TEXT, bold=(c == 1))
        elif ch and ch["status"] == "changed":
            for f in ch["fields"]:
                if f in headers:
                    cell = ws.cell(row=r, column=headers.index(f) + 2)
                    cell.fill = PatternFill("solid", fgColor=CHG_FILL)
                    cell.font = Font(name=FONT, size=10, color=CHG_TEXT, bold=True)
        if empty_col and str(row.get("empty")) == "Yes":
            cell = ws.cell(row=r, column=empty_col)
            cell.fill = PatternFill("solid", fgColor=EMPTY_FILL)
            cell.font = Font(name=FONT, size=10, bold=True, color=EMPTY_TEXT)
        r += 1
    autosize(ws)


def build_hba_tab(wb, arrays):
    ws = wb.create_sheet(title="All-HBA-WWNs")
    ws.sheet_view.showGridLines = False
    cols = ["serial", "port_id", "host_group_number", "host_group_name",
            "hba_wwn", "nickname", "source"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    t = ws.cell(row=1, column=1, value="HBA WWNs — logins (raidcom) + registered (module)")
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor=NAVY); t.alignment = LEFT
    ws.row_dimensions[1].height = 24
    hrow = 3
    for i, h in enumerate(cols, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(cols)); ws.freeze_panes = f"A{hrow+1}"
    allrecs = []
    for serial in sorted(arrays):
        for rec in all_hba_records(serial, arrays[serial]):
            allrecs.append((serial, rec))
    if not allrecs:
        ws.cell(row=hrow + 1, column=1,
                value="No HBA WWNs found — all host groups empty (no initiators registered or logged in).").font = \
            Font(name=FONT, italic=True, size=10, color=EMPTY_TEXT)
        autosize(ws); return
    r = hrow + 1
    for serial, rec in allrecs:
        vals = [serial, rec.get("port_id", ""), rec.get("host_group_number", ""),
                rec.get("host_group_name", ""), rec.get("hba_wwn", ""),
                rec.get("nickname", ""), rec.get("source", "")]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = LABEL_FONT if i == 1 else CELL_FONT
            cell.border = BORDER; cell.alignment = LEFT
            if (r - hrow) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        r += 1
    autosize(ws)


def pick_prev(args):
    if args.prev:
        return args.prev if os.path.isdir(args.prev) else None
    hist = sorted(glob.glob("exports/history/*"))
    return hist[-1] if hist else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default="exports/raw")
    ap.add_argument("--prev", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    if not os.path.isdir(args.raw):
        sys.exit(f"raw dir not found: {args.raw}")
    cur = load_raw(args.raw)
    if not cur:
        sys.exit("no array data found in raw JSON")
    for serial, a in cur.items():
        sysd = a.get("system", {})
        if isinstance(sysd, dict):
            a["_addr"] = sysd.get("controller_address", "")
        annotate_empty_hgs(serial, a)
    prev_dir = pick_prev(args)
    prev = load_raw(prev_dir) if prev_dir else {}
    for serial, a in prev.items():
        annotate_empty_hgs(serial, a)
    label = f"Compared against: {prev_dir}" if prev_dir else "No previous run found — baseline only (no diff)."
    changes, rows = diff_arrays(cur, prev)
    out = args.out or os.path.join(os.path.dirname(args.raw.rstrip("/")) or ".",
                                   f"vsp_facts_{dt.date.today().isoformat()}.xlsx")
    wb = Workbook()
    build_summary(wb, cur)
    build_changes_tab(wb, rows, label)
    for serial in sorted(cur):
        build_detail(wb, serial, cur[serial])
    for ft in FACT_TYPES:
        build_facttype_tab(wb, ft, cur, changes)
    build_hba_tab(wb, cur)
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb.save(out)
    print(f"  wrote {out}")
    print(f"  arrays: {', '.join(sorted(cur))}")
    print(f"  {label}; changes: {len(rows)}")


if __name__ == "__main__":
    main()
